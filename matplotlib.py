import re
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Color
from openpyxl.formatting.rule import CellIsRule

def should_skip_line(line):
    skip_prefixes = [
        "Result Gally Sheet",
        "College :",
        "College:",
        "College",
        "Course",
        "No of Papers",
        "No. of Papers",
        "Total Paper(s) :",
        "THIRUVALLUVAR",
        "SubCode",
        "UG",
        "Sub Code ",
        "Int",
        "NOTE",
        "UNIVERSITY",
        "ELIGIBLE",
        "NR",
        "*",
    ]
    return any(line.startswith(prefix) for prefix in skip_prefixes)

def normalize_mark(mark):
    if mark in ('---', '-'):
        return 0
    elif mark in ('WHI', 'WHE', 'AAA'):
        return mark  # Return mark as is for these string values
    else:
        try:
            return int(mark)
        except ValueError:
            return 0

def extract_semester_from_subject(subject_code):
    semester = ""
    for char in reversed(subject_code):
        if char.isdigit():
            semester = char
        elif semester:
            return semester
    return "Unknown"

def find_the_sem(exam_session):
    return [6, 4, 2] if exam_session == 'April/May' else [5, 3, 1]

def extract_student_data(lines, sems, previous_years):
    student_data_list = []
    register_number = ""
    student_name = ""
    subjects = {}
    idx = 0

    for line in lines:
        if should_skip_line(line):
            continue

        subject_matches = re.findall(r'(\d*[A-Z]+\s?\d+\s?[A-Z]*\s?\d*)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(PASS|RA|(?:A{3}|WHE|WHI))', line)
        if subject_matches:
            for match in subject_matches:
                subject_code = match[0].replace(" ", "")
                semester = extract_semester_from_subject(subject_code)
                internal_mark = normalize_mark(match[1])
                external_mark = normalize_mark(match[2])
                total_mark = normalize_mark(match[3])
                status = match[4] if match[4] != '---' else 'NA'

                adjusted_sems = sems[idx]

                if int(semester) == adjusted_sems:
                    subjects[subject_code] = total_mark

        match_register = re.search(r'(\d{3}\d{2}[A-Z]\d{5})\s+(.+)', line)
        if match_register:
            if student_name:
                student_data = {
                    "Register Number": register_number,
                    "Student Name": student_name,
                    **subjects,
                }
                student_data_list.append(student_data)
                subjects = {}

            register_number, student_name = match_register.groups()
            year_suffix = register_number[3:5]
            if int(year_suffix) in previous_years:
                idx = previous_years.index(int(year_suffix))

    if student_name:
        student_data = {
            "Register Number": register_number,
            "Student Name": student_name,
            **subjects,
        }
        student_data_list.append(student_data)

    return student_data_list

pdf_path = r"C:\pythonprogram\PDFCONVERT\MATHS.pdf"
with pdfplumber.open(pdf_path) as pdf:
    all_student_data = {}
    lines_buffer = []
    year = None
    exam_session_found = False
    exam_session = None
    current_year = None
    previous_years = []
    sems = []

    for page in pdf.pages:
        page_text = page.extract_text()
        lines = page_text.splitlines()

        for line in lines:
            if not exam_session_found:
                session_search = re.search(r'(Nov/Dec|April/May|Nov-Dec|April-May)\s+(\d{4})', line)
                if session_search:
                    exam_session, year_str = session_search.groups()
                    sems = find_the_sem(exam_session)
                    current_year = int(year_str)
                    previous_years = [(current_year - 2) % 100, (current_year - 1) % 100, (current_year) % 100]
                    exam_session_found = True
                continue

            match_register = re.search(r'(\d{3}\d{2}[A-Z]\d{5})\s+(.+)', line)
            if match_register:
                if lines_buffer:
                    student_data = extract_student_data(lines_buffer, sems, previous_years)
                    if year not in all_student_data and int(year) in previous_years:
                        all_student_data[year] = []
                    all_student_data[year].extend(student_data)
                    lines_buffer = []

                register_number, student_name = match_register.groups()
                year = register_number[3:5]

                if register_number and int(year) in previous_years:
                    lines_buffer.append(line)
            elif lines_buffer:
                lines_buffer.append(line)

    if lines_buffer:
        student_data = extract_student_data(lines_buffer, sems, previous_years)
        if year not in all_student_data and int(year) in previous_years:
            all_student_data[year] = []
        all_student_data[year].extend(student_data)

excel_file_name = "MATHS.xlsx"
dummy_df = pd.DataFrame({"Dummy": [0]})

if isinstance(current_year, int):
    years_to_process = [current_year]
else:
    years_to_process = [year for year in all_student_data.keys() if int(year) in current_year]

try:
    with pd.ExcelWriter(excel_file_name, engine="openpyxl") as writer:
        dummy_df.to_excel(writer, index=False, sheet_name="Dummy Sheet")

        for year, student_data in all_student_data.items():
            year_data_df = pd.DataFrame(student_data)

            status_list = []
            for index, row in year_data_df.iterrows():
                marks = row.drop(["Register Number", "Student Name"])
                if "AAA" in marks.values:
                    status_list.append("RA")
                elif (marks < 40).any():
                    status_list.append("RA")
                else:
                    status_list.append("PASS")
            year_data_df.insert(len(year_data_df.columns), "Status", status_list)

            year_sheet_name = f"{year}"
            year_data_df.to_excel(writer, index=False, sheet_name=year_sheet_name)

            sheet = writer.sheets[year_sheet_name]
            last_column_index = len(sheet['1']) + 1
            total_mark_col_letter = get_column_letter(last_column_index)
            
            sheet.cell(row=1, column=last_column_index, value="Total Mark")
            
            for row_index, row in year_data_df.iterrows():
                total_mark_formula = f"=SUM(C{row_index + 2}:{get_column_letter(last_column_index - 2)}{row_index + 2})"
                sheet.cell(row=row_index + 2, column=last_column_index, value=total_mark_formula)

            # Add count rows
            row_offset = len(year_data_df) + 3
            count_labels = ['Pass Count', 'Fail Count', 'AAA Count', 'WHE Count', 'WHI Count']
            b_font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            for label in count_labels:
                cell = sheet.cell(row=row_offset, column=2, value=label)
                cell.font = b_font
                cell.border = thin_border

                for col_index, subject_code in enumerate(year_data_df.columns[2:-1], start=3):
                    column_letter = get_column_letter(col_index)
                    if label == 'Pass Count':
                        formula = f'=COUNTIF({column_letter}2:{column_letter}{len(year_data_df)+1}, ">=40")'
                    elif label == 'Fail Count':
                        formula = f'=COUNTIF({column_letter}2:{column_letter}{len(year_data_df)+1}, "<40")'
                    elif label == 'AAA Count':
                        formula = f'=COUNTIF({column_letter}2:{column_letter}{len(year_data_df)+1}, "AAA")'
                    elif label == 'WHE Count':
                        formula = f'=COUNTIF({column_letter}2:{column_letter}{len(year_data_df)+1}, "WHE")'
                    elif label == 'WHI Count':
                        formula = f'=COUNTIF({column_letter}2:{column_letter}{len(year_data_df)+1}, "WHI")'
                    cell = sheet.cell(row=row_offset, column=col_index, value=formula)
                    cell.border = thin_border
                row_offset += 1

            # Apply border to the main data
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = thin_border

            # Apply conditional formatting to change font color to red for marks less than 40
            red_font = Font(color="FF0000")
            for col_index in range(3, last_column_index - 1):
                column_letter = get_column_letter(col_index)
                cell_range = f"{column_letter}2:{column_letter}{len(year_data_df) + 1}"
                sheet.conditional_formatting.add(cell_range, CellIsRule(operator="lessThan", formula=["40"], font=red_font))

            # Compute total marks for top three students using formulas
            for student in student_data:
                student["Total Mark"] = sum([v for k, v in student.items() if isinstance(v, int)])

            top_three_students = sorted(student_data, key=lambda x: x["Total Mark"], reverse=True)[:3]

            top_three_df = pd.DataFrame(top_three_students, columns=["Register Number", "Student Name", "Total Mark"])

            if len(student_data) > 0:
                top_three_start_row = len(student_data) + 10
            else:
                top_three_start_row = 3
            
            top_three_df.to_excel(writer, index=False, sheet_name=year_sheet_name, startrow=top_three_start_row)
        
        if "Dummy Sheet" in writer.book.sheetnames:
            writer.book.remove(writer.book["Dummy Sheet"])
    
    print(f"Success: Excel file '{excel_file_name}' created successfully.")
except Exception as e:
    print(f"Error while creating Excel file: {e}")
