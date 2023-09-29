import re
import pdfplumber
import pandas as pd
from openpyxl import Workbook

def should_skip_line(line):
    skip_prefixes = [
        "Result Gally Sheet",
        "College :",
        "College:",
        "College",
        "Course",
        "No of Papers",
        "No. of Papers",
        "Total Paper(s) :",
        "THIRUVALLUVAR",
        "SubCode",
        "UG",
        "Sub Code ",
        "Int",
        "NOTE",
        "UNIVERSITY",
        "ELIGIBLE",
        "NR",
        "*",
    ]
    return any(line.startswith(prefix) for prefix in skip_prefixes)

def normalize_mark(mark):
    if mark in ('---', '-'):
        return 'NA'
    elif mark in ('WHI', 'WHE'):
        return mark
    else:
        # Ensure marks are three digits
        return str(mark).zfill(3)

def extract_student_data(lines):
    student_data_list = []
    register_number = ""
    student_name = ""
    subjects = {}

    for line in lines:
        # Skip lines that should be removed
        if should_skip_line(line):
            continue

        subject_matches = re.findall(r'([A-Z]+\s?\d+\s?[A-Z]*\s?\d*)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(PASS|RA|(?:A{3}|WHE|WHI))', line)
        if subject_matches:
            for match in subject_matches:
                subject_code = match[0].replace(" ", "")
                internal_mark = normalize_mark(match[1])
                external_mark = normalize_mark(match[2])
                total_mark = normalize_mark(match[3])
                status = match[4] if match[4] != '---' else 'NA'

                # Add subject data to the dictionary for the current student
                subjects[subject_code] = {
                    "Subject Code": subject_code,
                    "Internal Mark": internal_mark,
                    "External Mark": external_mark,
                    "Total Mark": total_mark,
                    "Status": status
                }

        # Extract and format the register number and student name
        match_register = re.search(r'(\d{5}[A-Z](\d{2}))\s+(.+)', line)
        if match_register:
            # If we already have subjects, add the current student's data to the list
            if subjects:
                student_data_list.append({
                    "Register Number": register_number,
                    "Student Name": student_name,
                    "Subjects": subjects
                })
            register_number, year, student_name = match_register.groups()
            year = int(year)
            subjects = {}  # Reset subjects for the new student

    # Check if we have any remaining subjects for the last student
    if subjects:
        student_data_list.append({
            "Register Number": register_number,
            "Student Name": student_name,
            "Subjects": subjects
        })

    return student_data_list

# Open the PDF file
pdf_path = r"C:/pythonprogram/PDFCONVERT/5th.pdf"
with pdfplumber.open(pdf_path) as pdf:
    all_student_data = []
    lines_buffer = []

    for page in pdf.pages:
        page_text = page.extract_text()
        lines = page_text.splitlines()

        # Add lines to the buffer until a register number is found
        for line in lines:
            match_register = re.search(r'(\d{5}[A-Z](\d{2}))\s+(.+)', line)
            if match_register:
                if lines_buffer:
                    all_student_data.extend(extract_student_data(lines_buffer))
                    lines_buffer = []
                register_number, year, student_name = match_register.groups()
                year = int(year)
                lines_buffer.append(line)
            elif lines_buffer:
                lines_buffer.append(line)

    # Check if there are remaining lines in the buffer
    if lines_buffer:
        all_student_data.extend(extract_student_data(lines_buffer))

# Create an Excel workbook with a default empty sheet
wb = Workbook()
default_sheet = wb.active
default_sheet.title = "Default"

# Iterate through the student data and create sheets for each year
for student_data in all_student_data:
    register_number = student_data['Register Number']
    student_name = student_data['Student Name']
    year = student_data['Year']

    # Check if the sheet for the current year exists, if not, create one
    sheet = wb.get_sheet_by_name(str(year)) if str(year) in wb.sheetnames else wb.create_sheet(title=str(year))

    # Create column headers
    sheet.append(["Register Number", "Student Name"] + list(student_data['Subjects'].keys()))

    # Add student data to the sheet
    sheet.append([register_number, student_name] + [subject_data['Total Mark'] for subject_data in student_data['Subjects'].values()])

# Remove the default empty sheet
wb.remove(default_sheet)

# Save the workbook to an Excel file
excel_file = "student_data.xlsx"
wb.save(excel_file)

print(f"Excel file '{excel_file}' has been created with separate sheets for each year.")
