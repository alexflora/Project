import re
import pdfplumber
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

def should_skip_line(line):
    skip_prefixes = [
        "Result Gally Sheet",
        "College :",
        "College:",
        "College",
        "Course",
        "No of Papers",
        "No. of Papers",
        "Total Paper(s) :",
        "THIRUVALLUVAR",
        "SubCode",
        "UG",
        "Sub Code ",
        "Int",
        "NOTE",
        "UNIVERSITY",
        "ELIGIBLE",
        "NR",
        "*",
    ]
    return any(line.startswith(prefix) for prefix in skip_prefixes)

def normalize_mark(mark):
    if mark in ('---', '-'):
        return 0
    elif mark in ('WHI', 'WHE', 'AAA'):
        return mark  # Return 0 for these string values
    else:
        # Ensure marks are integers
        try:
            return int(mark)
        except ValueError:
            return 0
def extract_semester_from_subject(subject_code):
    semester = ""
    for char in reversed(subject_code):
        if char.isdigit():
            semester = char
        elif semester:
            # If a digit was found before encountering a character, return the semester
            return semester
    # If no digit was found, return a default value
    return "Unknown"

def find_the_sem(exam_session):
    return [6,4,2] if exam_session == 'April/May' else [5,3,1]
        
def extract_student_data(lines, sems, previous_years):
    student_data_list = []
    register_number = ""
    student_name = ""
    subjects = {}
    idx = 0  # Initialize index position to 0

    for line in lines:
        # Skip lines that should be removed
        if should_skip_line(line):
            continue

        subject_matches = re.findall(r'(\d*[A-Z]+\s?\d+\s?[A-Z]*\s?\d*)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(-{1,3}|\d+|A{3}|WHE|WHI)\s+(PASS|RA|(?:A{3}|WHE|WHI))', line)
        if subject_matches:
            for match in subject_matches:
                subject_code = match[0].replace(" ", "")
                semester = extract_semester_from_subject(subject_code)
                internal_mark = normalize_mark(match[1])
                external_mark = normalize_mark(match[2])
                total_mark = normalize_mark(match[3])
                status = match[4] if match[4] != '---' else 'NA'

                # Get adjusted semesters based on the current index position
                adjusted_sems = sems[idx]

                # Check if the semester of the subject matches the semesters to be included for the current year
                if int(semester) == adjusted_sems:
                    # Store the subject's total mark only if it belongs to the relevant semester
                    subjects[subject_code] = total_mark

        match_register = re.search(r'(\d{3}\d{2}[A-Z]\d{5})\s+(.+)', line)
        if match_register:
            if student_name:  # If there is student data, save it
                student_data = {
                    "Register Number": register_number,
                    "Student Name": student_name,
                    **subjects,  # Flatten subject data
                }
                student_data_list.append(student_data)
                subjects = {}  # Reset subjects for the next student

            register_number, student_name = match_register.groups()
            year_suffix = register_number[3:5]  # Extract year from the register number
            if int(year_suffix) in previous_years:
                idx = previous_years.index(int(year_suffix))  # Update index position for the current year
    if student_name:
        student_data = {
            "Register Number": register_number,
            "Student Name": student_name,
            **subjects,  # Flatten subject data
        }
        student_data_list.append(student_data)

    return student_data_list

pdf_path = r"C:\pythonprogram\PDFCONVERT\BCA.pdf"
with pdfplumber.open(pdf_path) as pdf:
    all_student_data = {}
    lines_buffer = []
    year = None
    exam_session_found=False
    exam_session = None
    current_year = None
    previous_years = []
    sems=[]
    previous_years = [str(year) for year in previous_years]
    
    for page in pdf.pages:
        page_text = page.extract_text()
        lines = page_text.splitlines()

        for line in lines:
            if not exam_session_found:
                session_search = re.search(r'(Nov/Dec|April/May|Nov-Dec|April-May)\s+(\d{4})', line)
                if session_search:
                    exam_session, year_str = session_search.groups()
                    sems = find_the_sem(exam_session)
                    current_year = int(year_str)
                    previous_years = [(current_year - 2)%100, (current_year - 1)%100,(current_year)%100]  # Calculate previous two years
                    exam_session_found = True
                continue
            match_register = re.search(r'(\d{3}\d{2}[A-Z]\d{5})\s+(.+)', line)
            if match_register:
                if lines_buffer:
                    student_data = extract_student_data(lines_buffer,sems,previous_years)
                    if year not in all_student_data and int(year) in previous_years:
                        all_student_data[year] = []
                    all_student_data[year].extend(student_data)
                    lines_buffer = []
                
                register_number, student_name = match_register.groups()
                year = register_number[3:5]

                # Filter and store only the register numbers that match the pattern "U09"
                if register_number and int(year) in previous_years:
                    lines_buffer.append(line)
            elif lines_buffer:
                lines_buffer.append(line)
    # Process the last student's data
    if lines_buffer:
        student_data = extract_student_data(lines_buffer,sems,previous_years)
        if year not in all_student_data and int(year) in previous_years:
            all_student_data[year] = []
        all_student_data[year].extend(student_data)
        

# Define the Excel file name
excel_file_name = "BCA.xlsx"

# Create a dummy DataFrame for a dummy sheet
dummy_df = pd.DataFrame({"Dummy": [0]})

if isinstance(current_year, int):
    years_to_process = [current_year]  # If current_year is an integer, create sheet for that year only
else:
    years_to_process = [year for year in all_student_data.keys() if int(year) in current_year]  # Process years in current_year

try:
    with pd.ExcelWriter(excel_file_name, engine="openpyxl") as writer:
        dummy_df.to_excel(writer, index=False, sheet_name="Dummy Sheet")  # Create a dummy sheet

        for year, student_data in all_student_data.items():
            # Create a DataFrame for all students for the year
            year_data_df = pd.DataFrame(student_data)

            # Calculate status for each student
            status_list = []
            for index, row in year_data_df.iterrows():
                marks = row.drop(["Register Number", "Student Name"])
                if "AAA" in marks.values:
                    status_list.append("RA")  # If AAA is present, set status to RA
                elif (marks < 40).any():
                    status_list.append("RA")  # If any mark is less than 40, set status to RA
                else:
                    status_list.append("PASS")  # Otherwise, set status to PASS
            year_data_df.insert(len(year_data_df.columns), "Status", status_list)  # Insert Status column before Total Mark column

            # Write all student data to a separate sheet for each year
            year_sheet_name = f"{year}"
            year_data_df.to_excel(writer, index=False, sheet_name=year_sheet_name)

            # Get the last row and column index
            sheet = writer.sheets[year_sheet_name]
            last_column_index = len(sheet['1'])  # Assuming header is in row 1

            # Add the "Total Mark" column header
            sheet.cell(row=1, column=last_column_index + 1, value="Total Mark")

            # Calculate total marks for each student dynamically and write to Excel sheet
            for row_index, row in year_data_df.iterrows():
                total_mark = 0
                for col_index, cell_value in enumerate(row.drop(["Register Number", "Student Name"]), start=3):
                    if isinstance(cell_value, int):
                        total_mark += cell_value
                    elif cell_value not in ('WHI', 'WHE', 'AAA'):
                        # Only add the value if it's not 'WHI', 'WHE', or 'AAA'
                        total_mark += 0  # Treat these string values as 0
                sheet.cell(row=row_index + 2, column=last_column_index + 1, value=total_mark)

            # Get the top three students for each year
            top_three_students = sorted(student_data, key=lambda x: sum([v for k, v in x.items() if k not in ("Register Number", "Student Name")]), reverse=True)[:3]

            # Create a DataFrame for the top three students
            top_three_df = pd.DataFrame(top_three_students, columns=["Register Number", "Student Name"] + [f"Subject {i+1}" for i in range(len(top_three_students[0])-2)])

            # Write the top three students to the same sheet for each year
            if len(student_data) > 0:
                top_three_start_row = len(student_data) + 3  # Place the top three students after 3 rows from the last student
            else:
                top_three_start_row = 3  # If no student data, start at row 3
            top_three_df.to_excel(writer, index=False, sheet_name=year_sheet_name, startrow=top_three_start_row)

        # Remove the dummy sheet
        if "Dummy Sheet" in writer.book.sheetnames:
            writer.book.remove(writer.book["Dummy Sheet"])

    print(f"Success: Excel file '{excel_file_name}' created successfully.")
except Exception as e:
    print(f"Error while creating Excel file: {e}")
